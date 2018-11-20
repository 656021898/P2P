# -*_ coding:utf-8 _*-
import openpyxl
from common.ReadConfig import ReadConfig


class DoExcel:

    @staticmethod
    def get_top_row(ExcelName, SheetName):
        top_row = {}
        wb = openpyxl.load_workbook(ExcelName)
        sheet = wb[SheetName]
        for i in range(1, (sheet.max_column + 1)):
            top_row[sheet.cell(1, i).value] = i
        return top_row, sheet

    def read_excel(self, ExcelReadName, SheetReadName):
        """获取excel内容并返回[{},{},{}]格式"""
        excel_list = []
        top_row_read, sheet = DoExcel.get_top_row(ExcelReadName, SheetReadName)
        for i in range(2, (sheet.max_row + 1)):
            each_dict = {}
            if (str(ReadConfig("module","module").readConfig()) == "全部") or (ReadConfig("module","id").readConfig() == "全部"):
                for item in top_row_read.keys():
                    each_dict[item] = sheet.cell(i, top_row_read[item]).value
                excel_list.append(each_dict)
            else:
                if sheet.cell(i,top_row_read["module"]).value in str(ReadConfig("module","module").readConfig()).split(",")\
                    or (str(sheet.cell(i,top_row_read["id"])) in str(ReadConfig("module","id").readConfig()).split(",")):
                    for item in top_row_read.keys():
                        each_dict[item] = sheet.cell(i, top_row_read[item]).value
                    excel_list.append(each_dict)
                else:
                    continue
        return excel_list

    def write_excel(self, ExcelWriteName, SheetWriteName, Excel_list=[]):
        top_row_write, aa = DoExcel.get_top_row(ExcelWriteName, SheetWriteName)
        wb = openpyxl.load_workbook(ExcelWriteName)
        sheet = wb[SheetWriteName]
        #     if sheet.cell(j, 1).value is None:用来判断表格第一列，单元格为空的最小行号
        for item in Excel_list:
            for i in item.keys():
                #用例的编号比行号小1，通过id把数据写入对应的单元格
                sheet.cell(row=(int(item["id"])+1), column=top_row_write[i], value=str(item[i]))
        wb.save(ExcelWriteName)
        wb.close()

    def updata_mobile(self, ExcelReadName, SheetReadName):
        pass

if __name__ == "__main__":
    pass
    # list1 = [{"id": 1, "TestResult": 44, "real_code": 55, "res_json": 66}]
    # DoExcel().write_excel("/Users/liqingju/Documents/python/My_project/test_data/testCase.xlsx", "testCaseRead", list1)
    # print(DoExcel().read_excel("/Users/liqingju/Documents/python/My_project/test_data/testCase.xlsx", "testCaseRead"))

    # print(DoExcel().read_excel("../test_data/testCase.xlsx","testCaseRead"))
    # print(str(ReadConfig("module","module").readConfig() == "全部") or ReadConfig("module","id").readConfig() == "全部")