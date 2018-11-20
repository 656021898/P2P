# -*_ coding:utf-8 _*-
import configparser
from common.GetPath import GetPath
from common.ReadConfig import ReadConfig
import os
import openpyxl


class GetMobile:
    wb = openpyxl.load_workbook(os.path.join(GetPath.get_path(),ReadConfig("Excel","ExcelReadName").readConfig()))
    sheet = wb[(ReadConfig("Excel","SheetMobile").readConfig())]
    mobile = sheet.cell(1, 1).value

    def getMobile(self,dict_data):
        # print(eval(dict_data["data"])["mobilephone"])
        # print(str(int(self.mobile)+1))
        dict1 = eval(dict_data["data"])
        dict1["mobilephone"] = str(int(self.mobile)+1)
        dict_data["data"] = str(dict1)
        self.sheet.cell(row=1, column=1, value=str(int(self.mobile)+1))
        self.wb.save(os.path.join(GetPath.get_path(),ReadConfig("Excel","ExcelReadName").readConfig()))
        self.wb.close()

        return dict_data

if __name__ == '__main__':
    print(GetMobile().getMobile({'id': '1', 'module': '注册', 'host': 'http://47.107.168.87:8080', 'url': '/futureloan/mvc/api/member/register', 'data': '{"mobilephone":"${tel}","pwd":"123456","regname":"江江"}', 'method': 'get', 'code': 10000, 'description': '正常注册', 'TestResult': '未通过', 'real_code': '20110', 'res_json': "{'status': 0, 'code': '20110', 'data': None, 'msg': '手机号码已被注册'}"}))

    # dict1 = {"data":'{"mobile":"123"}'}
    # dict2 = eval(dict1["data"])
    # dict2["mobile"] ="234"
    # dict1["data"] = str(dict2)
    # print(dict1)





